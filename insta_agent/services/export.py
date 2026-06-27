import io
from openpyxl import Workbook
from flask import send_file

from insta_agent.models import Contact, ActivityLog, SmsLog


def export_contacts_excel(contacts: list) -> io.BytesIO:
  wb = Workbook()
  ws = wb.active
  ws.title = "Contacts"
  ws.append(["ردیف", "نام کاربری", "شناسه IG", "نام", "تلفن", "ایمیل", "منبع", "تاریخ"])
  for i, c in enumerate(contacts, 1):
    ws.append([
      i,
      c.ig_username,
      c.ig_user_id,
      c.full_name,
      c.phone,
      c.email,
      c.source,
      c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
    ])
  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)
  return buf


def export_phones_excel(contacts: list) -> io.BytesIO:
  wb = Workbook()
  ws = wb.active
  ws.title = "Phones"
  ws.append(["ردیف", "شماره", "نام کاربری", "نام"])
  row = 1
  for c in contacts:
    if c.phone:
      ws.append([row, c.phone, c.ig_username, c.full_name])
      row += 1
  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)
  return buf


def export_activity_excel(logs: list) -> io.BytesIO:
  wb = Workbook()
  ws = wb.active
  ws.title = "Activity"
  ws.append(["ردیف", "نوع", "قانون", "کاربر", "اکشن", "وضعیت", "تاریخ"])
  for i, log in enumerate(logs, 1):
    ws.append([
      i, log.rule_type, log.rule_name,
      log.ig_username or log.ig_user_id,
      log.action, log.status,
      log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
    ])
  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)
  return buf
